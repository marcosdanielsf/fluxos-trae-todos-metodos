#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script Reparador Completo de Arquivos
Corrige problemas específicos encontrados nos arquivos
"""

import os
import sys
import json
import shutil
import hashlib
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import zipfile
import tempfile

# Bibliotecas para manipulação de arquivos
try:
    import pandas as pd
    from docx import Document
    import PyPDF2
    from pptx import Presentation
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.exceptions import InvalidFileException
    import fitz  # PyMuPDF
except ImportError as e:
    print(f"Erro ao importar bibliotecas: {e}")
    print("Execute: pip install pandas python-docx PyPDF2 python-pptx openpyxl PyMuPDF")
    sys.exit(1)

class ReparadorCompleto:
    def __init__(self, diretorio_base: str):
        self.diretorio_base = Path(diretorio_base)
        self.diretorio_consolidado = self.diretorio_base / "Consolidado"
        self.diretorio_logs = self.diretorio_consolidado / "logs"
        self.diretorio_reparados = self.diretorio_consolidado / "arquivos_reparados"
        self.diretorio_backup = self.diretorio_consolidado / "backup_originais"
        
        # Criar diretórios necessários
        self.diretorio_reparados.mkdir(exist_ok=True)
        self.diretorio_backup.mkdir(exist_ok=True)
        
        # Configurar logging
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.diretorio_logs / f"reparacao_completa_{self.timestamp}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
        # Estatísticas
        self.stats = {
            'arquivos_processados': 0,
            'arquivos_reparados': 0,
            'duplicados_removidos': 0,
            'extensoes_corrigidas': 0,
            'pdfs_reparados': 0,
            'xlsx_reparados': 0,
            'erros': 0
        }
        
        self.problemas_corrigidos = []
        self.erros_encontrados = []

    def calcular_checksum(self, arquivo: Path) -> str:
        """Calcula o checksum MD5 de um arquivo"""
        hash_md5 = hashlib.md5()
        try:
            with open(arquivo, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            self.logger.error(f"Erro ao calcular checksum de {arquivo}: {e}")
            return ""

    def criar_backup(self, arquivo: Path) -> bool:
        """Cria backup do arquivo original"""
        try:
            backup_path = self.diretorio_backup / arquivo.name
            shutil.copy2(arquivo, backup_path)
            self.logger.info(f"Backup criado: {backup_path}")
            return True
        except Exception as e:
            self.logger.error(f"Erro ao criar backup de {arquivo}: {e}")
            return False

    def reparar_pdf_sem_texto(self, arquivo: Path) -> bool:
        """Tenta reparar PDF sem texto extraível"""
        try:
            self.logger.info(f"Tentando reparar PDF: {arquivo}")
            
            # Criar backup
            if not self.criar_backup(arquivo):
                return False
            
            # Tentar abrir com PyMuPDF e recriar
            doc = fitz.open(arquivo)
            
            # Criar novo PDF
            novo_arquivo = self.diretorio_reparados / arquivo.name
            novo_doc = fitz.open()
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                # Tentar extrair texto
                text = page.get_text()
                
                if not text.strip():
                    # Se não há texto, tentar OCR ou manter como imagem
                    pix = page.get_pixmap()
                    img_data = pix.tobytes("png")
                    
                    # Criar nova página e inserir imagem
                    new_page = novo_doc.new_page(width=page.rect.width, height=page.rect.height)
                    new_page.insert_image(page.rect, stream=img_data)
                else:
                    # Copiar página com texto
                    novo_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
            
            novo_doc.save(novo_arquivo)
            novo_doc.close()
            doc.close()
            
            self.logger.info(f"PDF reparado salvo em: {novo_arquivo}")
            self.stats['pdfs_reparados'] += 1
            self.problemas_corrigidos.append({
                'arquivo': str(arquivo),
                'problema': 'PDF sem texto extraível',
                'solucao': 'PDF recriado com estrutura corrigida',
                'arquivo_reparado': str(novo_arquivo)
            })
            return True
            
        except Exception as e:
            self.logger.error(f"Erro ao reparar PDF {arquivo}: {e}")
            self.erros_encontrados.append({
                'arquivo': str(arquivo),
                'erro': str(e),
                'tipo': 'Reparo PDF'
            })
            return False

    def reparar_xlsx_com_erro(self, arquivo: Path) -> bool:
        """Tenta reparar arquivo XLSX com erro"""
        try:
            self.logger.info(f"Tentando reparar XLSX: {arquivo}")
            
            # Criar backup
            if not self.criar_backup(arquivo):
                return False
            
            # Tentar diferentes métodos de reparo
            metodos = [
                self._reparar_xlsx_openpyxl,
                self._reparar_xlsx_pandas,
                self._reparar_xlsx_zip
            ]
            
            for metodo in metodos:
                if metodo(arquivo):
                    return True
            
            return False
            
        except Exception as e:
            self.logger.error(f"Erro ao reparar XLSX {arquivo}: {e}")
            self.erros_encontrados.append({
                'arquivo': str(arquivo),
                'erro': str(e),
                'tipo': 'Reparo XLSX'
            })
            return False

    def _reparar_xlsx_openpyxl(self, arquivo: Path) -> bool:
        """Tenta reparar XLSX usando openpyxl"""
        try:
            # Tentar carregar e salvar novamente
            wb = load_workbook(arquivo, data_only=True)
            novo_arquivo = self.diretorio_reparados / arquivo.name
            wb.save(novo_arquivo)
            wb.close()
            
            self.logger.info(f"XLSX reparado com openpyxl: {novo_arquivo}")
            self.stats['xlsx_reparados'] += 1
            self.problemas_corrigidos.append({
                'arquivo': str(arquivo),
                'problema': 'Erro ao ler XLSX',
                'solucao': 'Arquivo recriado com openpyxl',
                'arquivo_reparado': str(novo_arquivo)
            })
            return True
            
        except Exception as e:
            self.logger.debug(f"Falha no reparo openpyxl para {arquivo}: {e}")
            return False

    def _reparar_xlsx_pandas(self, arquivo: Path) -> bool:
        """Tenta reparar XLSX usando pandas"""
        try:
            # Tentar ler com pandas e salvar novamente
            df = pd.read_excel(arquivo, sheet_name=None)
            novo_arquivo = self.diretorio_reparados / arquivo.name
            
            with pd.ExcelWriter(novo_arquivo, engine='openpyxl') as writer:
                for sheet_name, sheet_df in df.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            self.logger.info(f"XLSX reparado com pandas: {novo_arquivo}")
            self.stats['xlsx_reparados'] += 1
            self.problemas_corrigidos.append({
                'arquivo': str(arquivo),
                'problema': 'Erro ao ler XLSX',
                'solucao': 'Arquivo recriado com pandas',
                'arquivo_reparado': str(novo_arquivo)
            })
            return True
            
        except Exception as e:
            self.logger.debug(f"Falha no reparo pandas para {arquivo}: {e}")
            return False

    def _reparar_xlsx_zip(self, arquivo: Path) -> bool:
        """Tenta reparar XLSX tratando como ZIP"""
        try:
            # XLSX é um arquivo ZIP, tentar extrair e recriar
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir)
                
                # Extrair conteúdo
                with zipfile.ZipFile(arquivo, 'r') as zip_ref:
                    zip_ref.extractall(temp_path)
                
                # Recriar arquivo
                novo_arquivo = self.diretorio_reparados / arquivo.name
                with zipfile.ZipFile(novo_arquivo, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                    for file_path in temp_path.rglob('*'):
                        if file_path.is_file():
                            arcname = file_path.relative_to(temp_path)
                            zip_ref.write(file_path, arcname)
                
                # Verificar se o novo arquivo funciona
                wb = load_workbook(novo_arquivo)
                wb.close()
                
                self.logger.info(f"XLSX reparado como ZIP: {novo_arquivo}")
                self.stats['xlsx_reparados'] += 1
                self.problemas_corrigidos.append({
                    'arquivo': str(arquivo),
                    'problema': 'Erro ao ler XLSX',
                    'solucao': 'Arquivo recriado como ZIP',
                    'arquivo_reparado': str(novo_arquivo)
                })
                return True
                
        except Exception as e:
            self.logger.debug(f"Falha no reparo ZIP para {arquivo}: {e}")
            return False

    def remover_duplicados(self, duplicados: List[Dict]) -> int:
        """Remove arquivos duplicados"""
        removidos = 0
        
        for duplicado in duplicados:
            try:
                arquivo = Path(duplicado['arquivo'])
                if arquivo.exists():
                    # Criar backup antes de remover
                    self.criar_backup(arquivo)
                    
                    # Mover para pasta de duplicados
                    pasta_duplicados = self.diretorio_consolidado / "duplicados_removidos"
                    pasta_duplicados.mkdir(exist_ok=True)
                    
                    destino = pasta_duplicados / arquivo.name
                    shutil.move(arquivo, destino)
                    
                    self.logger.info(f"Duplicado removido: {arquivo} -> {destino}")
                    removidos += 1
                    
                    self.problemas_corrigidos.append({
                        'arquivo': str(arquivo),
                        'problema': 'Arquivo duplicado',
                        'solucao': 'Arquivo movido para pasta de duplicados',
                        'arquivo_movido': str(destino)
                    })
                    
            except Exception as e:
                self.logger.error(f"Erro ao remover duplicado {duplicado['arquivo']}: {e}")
                self.erros_encontrados.append({
                    'arquivo': duplicado['arquivo'],
                    'erro': str(e),
                    'tipo': 'Remoção duplicado'
                })
        
        self.stats['duplicados_removidos'] = removidos
        return removidos

    def corrigir_extensoes(self, arquivos_extensao_incorreta: List[Dict]) -> int:
        """Corrige extensões de arquivos incorretas"""
        corrigidos = 0
        
        for item in arquivos_extensao_incorreta:
            try:
                arquivo = Path(item['arquivo'])
                if not arquivo.exists():
                    continue
                
                # Detectar tipo real do arquivo
                tipo_real = self._detectar_tipo_arquivo(arquivo)
                if not tipo_real:
                    continue
                
                # Criar novo nome com extensão correta
                novo_nome = arquivo.stem + tipo_real
                novo_caminho = arquivo.parent / novo_nome
                
                # Evitar conflitos de nome
                contador = 1
                while novo_caminho.exists():
                    novo_nome = f"{arquivo.stem}_{contador}{tipo_real}"
                    novo_caminho = arquivo.parent / novo_nome
                    contador += 1
                
                # Criar backup
                self.criar_backup(arquivo)
                
                # Renomear arquivo
                arquivo.rename(novo_caminho)
                
                self.logger.info(f"Extensão corrigida: {arquivo} -> {novo_caminho}")
                corrigidos += 1
                
                self.problemas_corrigidos.append({
                    'arquivo': str(arquivo),
                    'problema': 'Extensão incorreta',
                    'solucao': f'Extensão corrigida para {tipo_real}',
                    'arquivo_corrigido': str(novo_caminho)
                })
                
            except Exception as e:
                self.logger.error(f"Erro ao corrigir extensão de {item['arquivo']}: {e}")
                self.erros_encontrados.append({
                    'arquivo': item['arquivo'],
                    'erro': str(e),
                    'tipo': 'Correção extensão'
                })
        
        self.stats['extensoes_corrigidas'] = corrigidos
        return corrigidos

    def _detectar_tipo_arquivo(self, arquivo: Path) -> Optional[str]:
        """Detecta o tipo real do arquivo baseado no conteúdo"""
        try:
            with open(arquivo, 'rb') as f:
                header = f.read(8)
            
            # Assinaturas de arquivos
            assinaturas = {
                b'\x50\x4B\x03\x04': '.zip',  # ZIP/Office
                b'\x50\x4B\x05\x06': '.zip',  # ZIP vazio
                b'\x50\x4B\x07\x08': '.zip',  # ZIP spanned
                b'\x25\x50\x44\x46': '.pdf',  # PDF
                b'\xFF\xD8\xFF': '.jpg',       # JPEG
                b'\x89\x50\x4E\x47': '.png',   # PNG
                b'\x47\x49\x46\x38': '.gif',   # GIF
            }
            
            for assinatura, extensao in assinaturas.items():
                if header.startswith(assinatura):
                    # Para arquivos Office, verificar mais especificamente
                    if extensao == '.zip':
                        return self._detectar_tipo_office(arquivo)
                    return extensao
            
            return None
            
        except Exception:
            return None

    def _detectar_tipo_office(self, arquivo: Path) -> Optional[str]:
        """Detecta tipo específico de arquivo Office"""
        try:
            with zipfile.ZipFile(arquivo, 'r') as zip_ref:
                files = zip_ref.namelist()
                
                if 'word/document.xml' in files:
                    return '.docx'
                elif 'xl/workbook.xml' in files:
                    return '.xlsx'
                elif 'ppt/presentation.xml' in files:
                    return '.pptx'
                else:
                    return '.zip'
                    
        except Exception:
            return '.zip'

    def processar_relatorio_problemas(self, arquivo_relatorio: Path) -> None:
        """Processa o relatório de problemas e executa reparos"""
        try:
            with open(arquivo_relatorio, 'r', encoding='utf-8') as f:
                relatorio = json.load(f)
            
            problemas = relatorio.get('problemas_encontrados', {})
            
            self.logger.info("Iniciando processo de reparo...")
            
            # Reparar PDFs sem texto
            pdfs_problematicos = problemas.get('arquivos_com_conteudo_suspeito', [])
            for item in pdfs_problematicos:
                if 'PDF sem texto extraível' in item.get('problema', ''):
                    arquivo = Path(item['arquivo'])
                    self.reparar_pdf_sem_texto(arquivo)
                elif 'Erro ao ler XLSX' in item.get('problema', ''):
                    arquivo = Path(item['arquivo'])
                    self.reparar_xlsx_com_erro(arquivo)
                
                self.stats['arquivos_processados'] += 1
            
            # Remover duplicados
            duplicados = problemas.get('arquivos_duplicados', [])
            if duplicados:
                self.logger.info(f"Removendo {len(duplicados)} arquivos duplicados...")
                self.remover_duplicados(duplicados)
            
            # Corrigir extensões
            extensoes_incorretas = problemas.get('arquivos_com_extensao_incorreta', [])
            if extensoes_incorretas:
                self.logger.info(f"Corrigindo {len(extensoes_incorretas)} extensões incorretas...")
                self.corrigir_extensoes(extensoes_incorretas)
            
            self.stats['arquivos_reparados'] = len(self.problemas_corrigidos)
            
        except Exception as e:
            self.logger.error(f"Erro ao processar relatório: {e}")
            self.erros_encontrados.append({
                'arquivo': str(arquivo_relatorio),
                'erro': str(e),
                'tipo': 'Processamento relatório'
            })

    def gerar_relatorio_final(self) -> str:
        """Gera relatório final do processo de reparo"""
        relatorio = {
            'timestamp': self.timestamp,
            'estatisticas': self.stats,
            'problemas_corrigidos': self.problemas_corrigidos,
            'erros_encontrados': self.erros_encontrados,
            'diretorios': {
                'arquivos_reparados': str(self.diretorio_reparados),
                'backup_originais': str(self.diretorio_backup),
                'logs': str(self.diretorio_logs)
            }
        }
        
        arquivo_relatorio = self.diretorio_logs / f"relatorio_reparo_final_{self.timestamp}.json"
        
        with open(arquivo_relatorio, 'w', encoding='utf-8') as f:
            json.dump(relatorio, f, indent=2, ensure_ascii=False)
        
        return str(arquivo_relatorio)

def main():
    # Configuração
    diretorio_base = "/Users/marcosdaniels/Downloads/5. FULL SALES 2"
    
    # Verificar se existe relatório de problemas
    diretorio_logs = Path(diretorio_base) / "Consolidado" / "logs"
    relatorios_problemas = list(diretorio_logs.glob("relatorio_problemas_especificos_*.json"))
    
    if not relatorios_problemas:
        print("❌ Nenhum relatório de problemas encontrado!")
        print("Execute primeiro o detector_problemas_especificos.py")
        return
    
    # Usar o relatório mais recente
    relatorio_mais_recente = max(relatorios_problemas, key=lambda x: x.stat().st_mtime)
    
    print("🔧 REPARADOR COMPLETO DE ARQUIVOS")
    print("=" * 60)
    print(f"📁 Diretório base: {diretorio_base}")
    print(f"📋 Relatório de problemas: {relatorio_mais_recente.name}")
    print("=" * 60)
    
    # Inicializar reparador
    reparador = ReparadorCompleto(diretorio_base)
    
    # Processar problemas
    reparador.processar_relatorio_problemas(relatorio_mais_recente)
    
    # Gerar relatório final
    relatorio_final = reparador.gerar_relatorio_final()
    
    # Exibir resultados
    print("\n" + "=" * 60)
    print("RELATÓRIO FINAL - REPARO DE ARQUIVOS")
    print("=" * 60)
    print(f"Arquivos processados: {reparador.stats['arquivos_processados']}")
    print(f"Arquivos reparados: {reparador.stats['arquivos_reparados']}")
    print(f"Duplicados removidos: {reparador.stats['duplicados_removidos']}")
    print(f"Extensões corrigidas: {reparador.stats['extensoes_corrigidas']}")
    print(f"PDFs reparados: {reparador.stats['pdfs_reparados']}")
    print(f"XLSX reparados: {reparador.stats['xlsx_reparados']}")
    print(f"Erros encontrados: {reparador.stats['erros']}")
    print("=" * 60)
    print(f"\n📊 Relatório completo salvo em: {relatorio_final}")
    print(f"🔧 Arquivos reparados em: {reparador.diretorio_reparados}")
    print(f"💾 Backups em: {reparador.diretorio_backup}")
    print("\nProcesso concluído! ✅")

if __name__ == "__main__":
    main()